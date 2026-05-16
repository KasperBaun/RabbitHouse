"""Restructure materialeliste.xlsx into per-variant sheets + summary.

Layout:
  summary                       — variant dropdowns + selected totals + comparison matrix
  fundament                     — shared
  konstruktion                  — shared
  tagkonstruktion_tagpap        — full tagpap roof BOM (rafters 45×95)
  tagkonstruktion_eternit_b7    — full eternit roof BOM (rafters 47×100, B7 plates)
  beklaedning_klink             — klink cladding BOM (incl. shared housewrap/insulation/mesh)
  beklaedning_board_on_board    — 1-på-2 cladding BOM (incl. shared bits)

Each data sheet uses a 9-column layout:
  A Kategori | B Vare | C Antal | D Enhed | E Pris/enh | F I alt | G Leverandør | H Noter | I Zone

Zone is one of: "Hus", "Yard", "Fælles". The house/yard split mirrors the
RenderHouse*/RenderYard* code split at X=RH_HOUSE_LEN (1500 mm). Items that
genuinely span both zones (continuous foundation ring, OSB deck whose plates
straddle hl, eternit waves that don't align with hl) stay "Fælles".

F column is a formula =C*E. SUM(F:F) gives the sheet total.
"""
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

XLSX = r'C:\dev\privat\rabbit-house\docs\materialeliste.xlsx'

# ---------------------------------------------------------------------------
# Colors — match the Office theme tints used in the original master sheet
# (theme 4/5/6/8 with tint 0.8 → light accent colors). We use direct RGB
# so it renders the same regardless of the user's installed Office theme.
# Each data sheet gets its own category color on column A.
# ---------------------------------------------------------------------------
COL_FUNDAMENT  = "D9E2F3"   # theme 4 tint 0.8 — light blue
COL_KONSTRUKT  = "FBE5D6"   # theme 5 tint 0.8 — peach
COL_TAG_TAGPAP = "EDEDED"   # theme 6 tint 0.8 — light gray  (preserved original)
COL_TAG_ETERN  = "FFF2CC"   # theme 7 tint 0.8 — light yellow
COL_CLAD_KLINK = "DEEBF7"   # theme 8 tint 0.8 — light cyan  (preserved original)
COL_CLAD_BOB   = "E2EFDA"   # theme 9 tint 0.8 — light green
COL_VOLIERE    = "F8CBAD"   # peach accent for the Voliere subsection

HEADER_BG       = "1F3864"   # dark navy — header row background
HEADER_FG       = "FFFFFF"   # white text on dark header
SUBHEADER_BG    = "8EA9DB"   # mid blue for section sub-headers in summary
TOTAL_BG        = "FFD966"   # gold for total cells
DELTA_POS_BG    = "FFE5D9"   # light pink for positive delta (extra cost)
DELTA_NEG_BG    = "E2EFDA"   # light green for negative delta (savings)
LINK_BLUE       = "0563C1"   # standard Excel hyperlink color

# Per-sheet color mapping (column A fill + main badge color)
SHEET_COLORS = {
    "fundament":                  COL_FUNDAMENT,
    "konstruktion":               COL_KONSTRUKT,
    "tagkonstruktion_tagpap":     COL_TAG_TAGPAP,
    "tagkonstruktion_eternit_b7": COL_TAG_ETERN,
    "beklaedning_klink":          COL_CLAD_KLINK,
    "beklaedning_board_on_board": COL_CLAD_BOB,
}

# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------
BOLD            = Font(bold=True)
HEADER_FONT     = Font(bold=True, color=HEADER_FG, size=11)
HEADER_FILL     = PatternFill("solid", fgColor=HEADER_BG)
SUBHEADER_FILL  = PatternFill("solid", fgColor=SUBHEADER_BG)
TOTAL_FILL      = PatternFill("solid", fgColor=TOTAL_BG)
DELTA_POS_FILL  = PatternFill("solid", fgColor=DELTA_POS_BG)
DELTA_NEG_FILL  = PatternFill("solid", fgColor=DELTA_NEG_BG)
LINK_FONT       = Font(color=LINK_BLUE, underline="single")
THIN            = Side(border_style="thin", color="BFBFBF")
THIN_BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---------------------------------------------------------------------------
# Data — preserved from the original sheet, split by category and variant.
# Each row: (kategori, vare, antal, enhed, pris_enh, leverandør, noter)
# ---------------------------------------------------------------------------

# Each row: (kategori, vare, antal, enhed, pris_enh, leverandør_tekst, noter, url, zone)
# URL may be None when no specific product link is known — leverandør then
# renders as plain text. EST_PRICE rows (estimates) carry "est." in noter so
# the user can spot which prices need verification. Zone is "Hus" | "Yard"
# | "Fælles" — the materialeliste sums correctly regardless, but the column
# lets the user filter cost per zone.

URL_MATPLADS_STABILGRUS = "https://materialepladsen.dk/produkter/grus/Stabilgrus0-32mm"
URL_MATPLADS_STOEBEMIX = "https://materialepladsen.dk/produkter/grus/Stoebemix0-16mm"
URL_JF_FUNDABLOK       = "https://www.jemogfix.dk/fundablok-graa-15-x-20-x-50-cm/4176/9033558/"
URL_JF_ARMERING        = "https://www.jemogfix.dk/kamstaal-8-mm-x-3-meter/4164/9042660/"
URL_JF_CEMENT          = "https://www.jemogfix.dk/aalborg-portland-basis-aalborg-cement-25-kg/4170/9814769/"
URL_JF_GEVINDSTANG     = "https://www.jemogfix.dk/nkt-fasteners-gevindstang-m10-x-1000-mm/5116/9711834/"
URL_JF_M10_MOETRIK     = "https://www.jemogfix.dk/nkt-fasteners-moetrik-elforzinket-m10-12-stk/5118/9513052/"

URL_JF_MURPAP_11       = "https://www.jemogfix.dk/phoenix-murpap-11-cm-x-20-meter/4210/9036209/"
URL_JF_REGLAR_45_3600  = "https://www.jemogfix.dk/impraegneret-reglar-45-x-95-x-3600-mm/4310/9057360/"
URL_JF_REGLAR_45_2400  = "https://www.jemogfix.dk/impraegneret-reglar-45-x-95-x-2400-mm/4310/9057358/"
URL_JF_REGLAR_47_3600  = "https://www.jemogfix.dk/reglar-47-x-100-x-3600-mm/4110/9057297/"
URL_JF_REGLAR_47_2400  = "https://www.jemogfix.dk/reglar-47-x-100-x-2400-mm/4110/9057295/"
URL_JF_REGLAR_47_3000  = "https://www.jemogfix.dk/reglar-47-x-100-x-3000-mm/4110/9057296/"

URL_JF_OSB             = "https://www.jemogfix.dk/osb-3-plade-tg4-18-mm-2397-x-600-mm/4138/9050213/"
URL_JF_STERN           = "https://www.jemogfix.dk/impraegneret-stern-mellem-21-x-120-x-3600-mm/4321/9057378/"
URL_JF_TAGPAP          = "https://www.jemogfix.dk/phoenix-selvbyggerpap-1-x-5-meter/4210/9021085/"
URL_JF_KLAEBEASFALT    = "https://www.jemogfix.dk/phoenix-klaebeasfalt-310-ml/4211/9014391/"
URL_JF_MURPAP_15       = "https://www.jemogfix.dk/phoenix-murpap-15-cm-x-20-meter/4210/9060998/"
URL_JF_ALU_TAGFOD      = "https://www.jemogfix.dk/tagfod-sort-aluminium-55-x-80-x-1000-mm/4210/9058507/"
URL_JF_STERNKAPSEL     = "https://www.jemogfix.dk/sternkapsel-lige-sort-35-x-25-mm-x-100-cm/4210/9062741/"
URL_JF_TAGRENDE        = "https://www.jemogfix.dk/staaltagrende-med-silver-metallic-4-meter/4203/9000602/"
URL_JF_KONSOLJERN      = "https://www.jemogfix.dk/konsoljern-staal-90-grader/4203/9000608/"
URL_JF_SAMLESTYKKE     = "https://www.jemogfix.dk/samlestykke-staal/4203/9000604/"
URL_JF_ENDEBUND        = "https://www.jemogfix.dk/rodena-endebund-staal-hoejrevenstre/4203/9000603/"
URL_JF_BLADSAMLER      = "https://www.jemogfix.dk/bladsamler-i-plast-75-82-mm/4202/9711493/"
URL_JF_NEDLOEBSROER    = "https://www.jemogfix.dk/nedloebsroer-staal-3-m/4203/9000611/"
URL_JF_NEDLOEBSBOEJ    = "https://www.jemogfix.dk/rodena-boejning-70-gr-staal/4203/9000612/"
URL_JF_TUDSTYKKE       = "https://www.jemogfix.dk/tudstykke-staal-oe75-mm/4203/9000605/"
URL_JF_ALUSOEM         = "https://www.jemogfix.dk/alusoem-2-6-x-25-mm-100-stk/4210/9038151/"

URL_JF_KLINK           = "https://www.jemogfix.dk/klinkbeklaedning-sortmalet/4115/9023155/"
URL_JF_VINDPAP         = "https://www.jemogfix.dk/phoenix-isoleringspap-rf-650-1-0-x-20-meter/4210/9358237/"
URL_JF_KLEMMELISTE     = "https://www.jemogfix.dk/impraegneret-forskalling-fyr-25-x-50-x-4200-mm/4310/9063674/"
URL_JF_HJORNETRIM_45   = "https://www.jemogfix.dk/reglar-47-x-50-x-3000-mm/4110/9052623/"
URL_JF_ISOLERING       = "https://www.jemogfix.dk/glasuld-isover-basic-ruller-95-mm/4180/9039070/"

# Eternit-specific URLs (from docs/tagkonstruktion_eternit.md)
URL_104_LAEGTE         = "https://www.10-4.dk/varer/byggematerialer/trae/laegter/38x73mm-taglaegte-c18-1731038073"
URL_XLBYG_STERN        = "https://www.xl-byg.dk/produkt/xl-byg-ru-sternbraet-over-trykimpraegneret-25-x-125-x-3600-mm-1143025122-0360"
URL_BYGXTRA_B7         = "https://www.bygxtra.dk/produkter/swisspearl-boelgeplade-b7-fk-i-sortblaa-1100x570-mm-1617012"
URL_DAVIDSEN_SKRUE     = "https://www.davidsen.dk/swisspearl-100-tagskrue-m-taetningsskive-6x100-mm-graa-100-stk-c-id525046-p-51601506116"
URL_WOOD_VINKEL        = "https://wood-online.dk/shop/paslode-vinkelbeslag-542p.html"

# Board-on-board URLs
URL_AROS_1PA2          = "https://arossavvaerk.dk/vare/1-paa-2-beklaedning-svensk-gran/"


# Foundation split: house owns perimeter [0..hl] front+back + V3 left + V5
# partition cross. Yard owns perimeter [hl..ll] front+back + V4 right.
# In COMBINED build, V5 is shared (house renders it). Linear-meter share:
# house ~8 m (41 %), yard ~11.5 m (59 %).
FUNDAMENT = [
    ("Fundament", "Stabilgrus 0-32 mm (hus)",       420, "kg",  0.55,  "materialepladsen.dk", "Under hus-strips (~0,25 m³)",                 URL_MATPLADS_STABILGRUS, "Hus"),
    ("Fundament", "Stabilgrus 0-32 mm (yard)",      605, "kg",  0.55,  "materialepladsen.dk", "Under yard-strips (~0,35 m³)",                URL_MATPLADS_STABILGRUS, "Yard"),
    ("Fundament", "Fundablok 50x20x15 cm (hus)",     70, "stk", 15.75, "jemogfix.dk",         "1500x2500 ring, 4 skifter halvstensforbandt — 64 + buffer", URL_JF_FUNDABLOK, "Hus"),
    ("Fundament", "Fundablok 50x20x15 cm (yard)",    75, "stk", 15.75, "jemogfix.dk",         "3-sidet ring (V5 = hus), 3 skifter halvstensforbandt — 69 + buffer", URL_JF_FUNDABLOK, "Yard"),
    ("Fundament", "Kamstål Ø10 mm × 3 m (hus)",      14, "stk", 45.00, "jemogfix.dk",         "est. — 2 bund + 2 top + lodrette",            URL_JF_ARMERING,         "Hus"),
    ("Fundament", "Kamstål Ø10 mm × 3 m (yard)",     18, "stk", 45.00, "jemogfix.dk",         "est. — 2 bund + 2 top + lodrette",            URL_JF_ARMERING,         "Yard"),
    ("Fundament", "Cement 25 kg (hus)",               8, "stk", 75.00, "jemogfix.dk",         "1:4 blanding, ~240 L beton",                  URL_JF_CEMENT,           "Hus"),
    ("Fundament", "Cement 25 kg (yard)",              9, "stk", 75.00, "jemogfix.dk",         "1:4 blanding, ~260 L beton",                  URL_JF_CEMENT,           "Yard"),
    ("Fundament", "Støbemix 0-16 mm (hus)",         500, "kg",  0.66,  "materialepladsen.dk", "ca. 0,30 m³",                                 URL_MATPLADS_STOEBEMIX,  "Hus"),
    ("Fundament", "Støbemix 0-16 mm (yard)",        700, "kg",  0.66,  "materialepladsen.dk", "ca. 0,42 m³",                                 URL_MATPLADS_STOEBEMIX,  "Yard"),
    ("Fundament", "Ankerskrue M10 × 120 mm (hus)",    6, "stk", 14.95, "jemogfix.dk",         "est. — bundrem-fast: 1 på V1, 1 på V2, 2 på V3, 2 på V5", URL_JF_GEVINDSTANG, "Hus"),
    ("Fundament", "Ankerskrue M10 × 120 mm (yard)",  10, "stk", 14.95, "jemogfix.dk",         "est. — 4 på V1, 4 på V2, 2 på V4",            URL_JF_GEVINDSTANG,      "Yard"),
    ("Fundament", "M10 møtrikker + skiver (12-pak)",  2, "pk.", 28.95, "jemogfix.dk",         "Fastgør bundrem til ankerskrue ovenfra (1 pak pr. zone)", URL_JF_M10_MOETRIK, "Fælles"),
]

# Konstruktion split: V3 + V5 (partition) + V1[0..hl] + V2[0..hl] = Hus.
# V4 + V1[hl..ll] + V2[hl..ll] = Yard. The two long 3600 mm bundrem/toprem
# pieces on V1+V2 straddle hl and stay "Fælles".
KONSTRUKTION = [
    ("Konstruktion", "Murpap (DPC)",                       1,  "stk", 99.00, "jemogfix.dk", "11 cm × 20 m — sokkel-isolering, hele perimeter + tværvæg", URL_JF_MURPAP_11,     "Fælles"),
    ("Konstruktion", "Bundrem 45×95×3600 PT NTR-AB",        2,  "stk", 71.10, "jemogfix.dk", "V1 + V2 lange stykker — straddler hl ved splejs",            URL_JF_REGLAR_45_3600, "Fælles"),
    ("Konstruktion", "Bundrem 45×95×2400 PT NTR-AB (hus)",  2,  "stk", 47.40, "jemogfix.dk", "V3 + V5 bundrem",                                            URL_JF_REGLAR_45_2400, "Hus"),
    ("Konstruktion", "Bundrem 45×95×2400 PT NTR-AB (yard)", 3,  "stk", 47.40, "jemogfix.dk", "V1[2] + V2[2] + V4 bundrem",                                 URL_JF_REGLAR_45_2400, "Yard"),
    ("Konstruktion", "Toprem 47×100×3600 C24 gran",         2,  "stk", 63.90, "jemogfix.dk", "V1 + V2 lange stykker — straddler hl",                       URL_JF_REGLAR_47_3600, "Fælles"),
    ("Konstruktion", "Toprem 47×100×3600 C24 gran (yard)",  2,  "stk", 63.90, "jemogfix.dk", "V1[2] + V2[2] yard-segment toprem",                          URL_JF_REGLAR_47_3600, "Yard"),
    ("Konstruktion", "Reglar 47×100×2400 C24 gran (hus)",  22,  "stk", 42.60, "jemogfix.dk", "Studs V1[0..hl]+V2[0..hl]+V3+V5 + skrå toprem V3+V5 + headers/cripples", URL_JF_REGLAR_47_2400, "Hus"),
    ("Konstruktion", "Reglar 47×100×2400 C24 gran (yard)", 20,  "stk", 42.60, "jemogfix.dk", "Studs V1[hl..ll]+V2[hl..ll]+V4 + skrå toprem V4 + yard-dør jamber", URL_JF_REGLAR_47_2400, "Yard"),
]

# Tagpap variant. House owns roof segment X<=hl + left side; yard owns
# X>=hl + right side. OSB plates tile across X — plate 1 fits within house
# segment, plates 2-3 in yard, so house 1 plate × 5 rows = 5, yard 2 × 5 = 10.
TAG_TAGPAP = [
    ("Træværk",        "Spær 47×100×3000 C24 gran (hus)",            4, "stk",   53.25, "jemogfix.dk",    "X=-220, 0, 600, 1200 (incl. vindskede V)", URL_JF_REGLAR_47_3000, "Hus"),
    ("Træværk",        "Spær 47×100×3000 C24 gran (yard)",           9, "stk",   53.25, "jemogfix.dk",    "X=1800..5400 + V4 gable + vindskede H",    URL_JF_REGLAR_47_3000, "Yard"),
    ("Træværk",        "OSB-3 TG4 18 mm 2397×600 mm (hus)",          5, "stk",  129.00, "jemogfix.dk",    "Plade 1 i hver række — dækker [-220..2177]", URL_JF_OSB,           "Hus"),
    ("Træværk",        "OSB-3 TG4 18 mm 2397×600 mm (yard)",        10, "stk",  129.00, "jemogfix.dk",    "Plade 2+3 i hver række — dækker [2177..6220]",URL_JF_OSB,          "Yard"),
    ("Træværk",        "Sternbræt imp. 21×120×3600 gran (hus)",      2, "stk",   92.70, "jemogfix.dk",    "Venstre side + del af front/bag (~6,4 m)",  URL_JF_STERN,          "Hus"),
    ("Træværk",        "Sternbræt imp. 21×120×3600 gran (yard)",     4, "stk",   92.70, "jemogfix.dk",    "Højre side + resten af front/bag (~12,4 m)", URL_JF_STERN,         "Yard"),
    ("Tagdækning",     "Phoenix Selvbyggerpap 1×5 m (hus)",          2, "stk",  579.00, "jemogfix.dk",    "~5 m² hus + spild",                         URL_JF_TAGPAP,         "Hus"),
    ("Tagdækning",     "Phoenix Selvbyggerpap 1×5 m (yard)",         4, "stk",  579.00, "jemogfix.dk",    "~14 m² yard + spild",                       URL_JF_TAGPAP,         "Yard"),
    ("Tagdækning",     "Phoenix klæbeasfalt 310 ml (hus)",           1, "stk",   69.95, "jemogfix.dk",    "",                                          URL_JF_KLAEBEASFALT,   "Hus"),
    ("Tagdækning",     "Phoenix klæbeasfalt 310 ml (yard)",          2, "stk",   69.95, "jemogfix.dk",    "",                                          URL_JF_KLAEBEASFALT,   "Yard"),
    ("Tagdækning",     "Phoenix murpap til tagfod 15 cm × 20 m",     1, "stk",  149.00, "jemogfix.dk",    "Én rulle dækker begge zoner",               URL_JF_MURPAP_15,      "Fælles"),
    ("Aluinddækning",  "Alu tagfod sort 55×80×1000 mm (hus)",        2, "stk",   39.75, "jemogfix.dk",    "Front+bag hus-segment + ende-trim",         URL_JF_ALU_TAGFOD,     "Hus"),
    ("Aluinddækning",  "Alu tagfod sort 55×80×1000 mm (yard)",       5, "stk",   39.75, "jemogfix.dk",    "Front+bag yard-segment + ende-trim",        URL_JF_ALU_TAGFOD,     "Yard"),
    ("Aluinddækning",  "Alu sternkapsel sort 35×25×1000 mm (hus)",   3, "stk",   59.95, "jemogfix.dk",    "Venstre side + front+bag hus",              URL_JF_STERNKAPSEL,    "Hus"),
    ("Aluinddækning",  "Alu sternkapsel sort 35×25×1000 mm (yard)", 10, "stk",   59.95, "jemogfix.dk",    "Højre side + front+bag yard",               URL_JF_STERNKAPSEL,    "Yard"),
    ("Beslag",         "Paslode vinkelbeslag 90×90×40 (20-pak) (hus)",1,"pak",   None,  "wood-online.dk", "TBD pris — ~14 stk på hus-spær",            URL_WOOD_VINKEL,       "Hus"),
    ("Beslag",         "Paslode vinkelbeslag 90×90×40 (20-pak) (yard)",2,"pak",  None,  "wood-online.dk", "TBD pris — ~34 stk på yard-spær",           URL_WOOD_VINKEL,       "Yard"),
    ("Beslag",         "OSB/spær-skruer 5×80 mm",                  250, "stk",   None,  "jemogfix.dk",    "TBD pris — én pakke dækker begge",          None,                  "Fælles"),
    ("Beslag",         "Rustfri tagskruer m. EPDM-pakning",       None, "stk",   None,  "jemogfix.dk",    "TBD antal + pris",                          None,                  "Fælles"),
    ("Beslag",         "Galvaniseret tagpapsøm 2,6×25 mm (100-pak)", 1, "pk.",   69.95, "jemogfix.dk",    "Alusøm — én pakke dækker begge",            URL_JF_ALUSOEM,        "Fælles"),
]

# Eternit-B7 variant. hl=1500 and RH_OH_SIDE=220 give 1720 mm = 10×B7_PITCH,
# so wave phase aligns at the partition line — house+yard plates tile
# seamlessly. Per row: house needs 2 plates × 7 rows ≈ 14; yard 5 × 7 ≈ 35.
# Plus 5 % spild. Battens split per row: house 1 stick (1720 < 4200) +
# yard 2 sticks (4720 > 4200) per row × 6 batten rows.
TAG_ETERNIT = [
    ("Træværk",        "Spær 47×100×3000 C24 gran (hus)",            4, "stk",   85.00, "Stark",          "est. — X<=1200 (incl. vindskede V)",        URL_JF_REGLAR_47_3000, "Hus"),
    ("Træværk",        "Spær 47×100×3000 C24 gran (yard)",           9, "stk",   85.00, "Stark",          "est. — X>=1800",                            URL_JF_REGLAR_47_3000, "Yard"),
    ("Træværk",        "C18 taglægte 38×73×4200 mm (hus)",           6, "stk",   52.50, "10-4.dk",        "6 batten-rækker × 1 stk hus-segment",       URL_104_LAEGTE,        "Hus"),
    ("Træværk",        "C18 taglægte 38×73×4200 mm (yard)",         12, "stk",   52.50, "10-4.dk",        "6 batten-rækker × 2 stk yard-segment",      URL_104_LAEGTE,        "Yard"),
    ("Træværk",        "Imp. sternbræt 25×125×3600 gran (hus)",      2, "stk",   82.60, "xl-byg.dk",      "Venstre side + del af front/bag (~6,4 m)",  URL_XLBYG_STERN,       "Hus"),
    ("Træværk",        "Imp. sternbræt 25×125×3600 gran (yard)",     4, "stk",   82.60, "xl-byg.dk",      "Højre side + resten af front/bag (~12,4 m)", URL_XLBYG_STERN,      "Yard"),
    ("Tagdækning",     "Swisspearl B7 bølgeplade FK 1100×570 sortblå (hus)", 15, "stk", 95.00, "bygxtra.dk", "2 plader × 7 rækker + 5 % spild",          URL_BYGXTRA_B7,        "Hus"),
    ("Tagdækning",     "Swisspearl B7 bølgeplade FK 1100×570 sortblå (yard)",37, "stk", 95.00, "bygxtra.dk", "5 plader × 7 rækker + 5 % spild",          URL_BYGXTRA_B7,        "Yard"),
    ("Tagdækning",     "Swisspearl 100 Tagskrue 6×100 (100-pak) (hus)", 1, "pak", 425.00, "davidsen.dk",    "2 pr. plade × 14 + spild",                  URL_DAVIDSEN_SKRUE,    "Hus"),
    ("Tagdækning",     "Swisspearl 100 Tagskrue 6×100 (100-pak) (yard)",2, "pak", 425.00, "davidsen.dk",    "2 pr. plade × 35 + spild",                  URL_DAVIDSEN_SKRUE,    "Yard"),
    ("Tagdækning",     "Swisspearl PVC skumstrimmel 4,5 mm",         1, "rl.",   None,  "Cembrit",        "TBD pris — overlap-tætning, én rulle deler", None,                 "Fælles"),
    ("Beslag",         "Paslode vinkelbeslag 90×90×40 (20-pak) (hus)",1,"pak",   None,  "wood-online.dk", "TBD pris — ~14 stk på hus-spær",            URL_WOOD_VINKEL,       "Hus"),
    ("Beslag",         "Paslode vinkelbeslag 90×90×40 (20-pak) (yard)",2,"pak",  None,  "wood-online.dk", "TBD pris — ~34 stk på yard-spær",           URL_WOOD_VINKEL,       "Yard"),
    ("Beslag",         "Lægtesøm varmforzinket 2,8×60",            None, "stk",   None,  "jemogfix.dk",    "TBD antal + pris",                          None,                  "Fælles"),
]

# Klink variant. Voliernet + klamper are on the YARD walls (front/back/right);
# klink + housewrap + battens + corner trim + isolering are on the 4 HOUSE
# walls only.
CLAD_KLINK = [
    ("Voliere",    "Voliernet welded 13×13 mm",                         2, "stk", 1350.00, "(ukendt)",    "1×10 m + 1×25 m — yard front+back+right",        None,                "Yard"),
    ("Voliere",    "Klamper til voliernet",                          None, "stk",  None,   "jemogfix.dk", "TBD antal + pris",                                 None,                "Yard"),
    ("Beklædning", "Klinkbeklædning sortmalet gran 25×125×4200 (pak)",   6, "pak",  497.70, "jemogfix.dk", "Frøslev klink; 4 hus-vægge",                       URL_JF_KLINK,        "Hus"),
    ("Beklædning", "Vindpap (housewrap)",                                1, "rl.",  299.00, "jemogfix.dk", "20 m² pr. rulle; behov ~17,5 m² på 4 hus-vægge",   URL_JF_VINDPAP,      "Hus"),
    ("Beklædning", "Klemmelister 25×50×4200 mm",                       10, "stk",   28.35, "jemogfix.dk", "Lodrette afstandslister c/c 600 på hus-vægge",     URL_JF_KLEMMELISTE,  "Hus"),
    ("Beklædning", "Hjørnetrim 45×45×2400 mm",                          4, "stk",   29.25, "jemogfix.dk", "4 udvendige hus-hjørner",                          URL_JF_HJORNETRIM_45,"Hus"),
    ("Beklædning", "Isolering mineraluld 95 mm",                       20, "m²",    29.61, "jemogfix.dk", "I hus-stud-rum (4 vægge)",                          URL_JF_ISOLERING,    "Hus"),
]

# Board-on-board variant. Same zone split as klink — voliernet yard, alt
# andet hus.
CLAD_BOB = [
    ("Voliere",    "Voliernet welded 13×13 mm",                         2, "stk", 1350.00, "(ukendt)",         "1×10 m + 1×25 m — yard front+back+right",         None,                "Yard"),
    ("Voliere",    "Klamper til voliernet",                          None, "stk",  None,   "jemogfix.dk",      "TBD antal + pris",                                None,                "Yard"),
    ("Beklædning", "1-på-2 svensk gran 25×150 savskåret",              16, "m²",   160.00, "arossavvaerk.dk",  "Aros tilbudspris; normal 225 kr/m² — hus-vægge",  URL_AROS_1PA2,       "Hus"),
    ("Beklædning", "Træolie/maling til naturligt træ",                  3, "L",     None,  "(vælg selv)",      "TBD — Gori/Cuprinol/Pinotex; klink er sortmalet", None,                "Hus"),
    ("Beklædning", "Vindpap (housewrap)",                                1, "rl.",  299.00, "jemogfix.dk",      "Samme som klink — hus-vægge",                     URL_JF_VINDPAP,      "Hus"),
    ("Beklædning", "Klemmelister 25×50×4200 mm",                       10, "stk",   28.35, "jemogfix.dk",      "Vandrette afstandslister c/c 600 på hus-vægge",   URL_JF_KLEMMELISTE,  "Hus"),
    ("Beklædning", "Hjørnetrim 70×70×2400 trykimp.",                    4, "stk",   None,  "jemogfix.dk",      "TBD — større end klink (bob stickout 50 mm)",     None,                "Hus"),
    ("Beklædning", "Isolering mineraluld 95 mm",                       20, "m²",    29.61, "jemogfix.dk",      "Samme som klink — hus-stud-rum",                  URL_JF_ISOLERING,    "Hus"),
    ("Beklædning", "Rustfri A4 facadeskruer 4×60",                   None, "pk.",   None,  "jemogfix.dk",      "TBD antal + pris — flere end klink (2 lag)",      None,                "Hus"),
]

# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------
HEADERS = ["Kategori", "Vare", "Antal", "Enhed", "Pris pr. enhed (DKK)",
           "I alt (DKK)", "Leverandør", "Noter", "Zone"]
COL_WIDTHS = [18, 50, 8, 8, 16, 14, 24, 60, 10]

ZONE_COLORS = {
    "Hus":    "FBE5D6",  # peach — matches konstruktion accent
    "Yard":   "E2EFDA",  # light green — matches yard mesh / open-air feel
    "Fælles": "F2F2F2",  # neutral light gray
}

def write_header(ws):
    for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

def write_rows(ws, rows, sheet_color, start_row=2):
    """rows = list of 9-tuples:
       (kat, vare, antal, enhed, pris, lev, noter, url, zone).
    sheet_color = hex (no #) for column A fill on this sheet."""
    cat_fill = PatternFill("solid", fgColor=sheet_color)
    voliere_fill = PatternFill("solid", fgColor=COL_VOLIERE)
    r = start_row
    last_kategori = None
    for row in rows:
        kat, vare, antal, enhed, pris, lev, noter, url, zone = row
        if kat != last_kategori and last_kategori is not None:
            r += 1   # blank separator

        # Column A — category badge (colored fill, bold first occurrence)
        a = ws.cell(row=r, column=1, value=kat)
        a.fill = voliere_fill if kat == "Voliere" else cat_fill
        a.font = Font(bold=(kat != last_kategori))

        ws.cell(row=r, column=2, value=vare)
        ws.cell(row=r, column=3, value=antal)
        ws.cell(row=r, column=4, value=enhed).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=5, value=pris)
        # "I alt" formula only when both antal+pris are numeric; else blank
        if isinstance(antal, (int, float)) and isinstance(pris, (int, float)):
            ws.cell(row=r, column=6, value=f"=C{r}*E{r}")
        else:
            ws.cell(row=r, column=6, value=None)

        lev_cell = ws.cell(row=r, column=7, value=lev)
        if url:
            lev_cell.hyperlink = url
            lev_cell.font = LINK_FONT

        ws.cell(row=r, column=8, value=noter).alignment = Alignment(wrap_text=False, vertical="center")

        z = ws.cell(row=r, column=9, value=zone)
        z.alignment = Alignment(horizontal="center")
        z.fill = PatternFill("solid", fgColor=ZONE_COLORS.get(zone, "FFFFFF"))
        z.font = BOLD

        for c in range(1, 10):
            ws.cell(row=r, column=c).border = THIN_BORDER

        last_kategori = kat
        r += 1
    return r   # next free row

def write_total(ws, last_row, sheet_color):
    """TOTAL row + per-zone subtotals (Hus / Yard / Fælles)."""
    cat_fill = PatternFill("solid", fgColor=sheet_color)

    r = last_row + 1
    ws.cell(row=r, column=1, value="").fill = cat_fill
    label = ws.cell(row=r, column=5, value="TOTAL")
    label.font = BOLD
    label.fill = TOTAL_FILL
    label.alignment = Alignment(horizontal="right")
    total_cell = ws.cell(row=r, column=8, value=f"=SUM(F2:F{last_row-1})")
    total_cell.font = Font(bold=True, size=11)
    total_cell.fill = TOTAL_FILL
    total_cell.number_format = '#,##0.00 "kr"'
    for c in range(1, 10):
        ws.cell(row=r, column=c).border = THIN_BORDER

    # Per-zone subtotals using SUMIF on column I.
    for i, zone in enumerate(("Hus", "Yard", "Fælles")):
        rr = r + 1 + i
        z_label = ws.cell(row=rr, column=5, value=f"heraf {zone}")
        z_label.font = BOLD
        z_label.alignment = Alignment(horizontal="right")
        z_label.fill = PatternFill("solid", fgColor=ZONE_COLORS[zone])
        z_total = ws.cell(row=rr, column=8,
                          value=f'=SUMIF(I2:I{last_row-1},"{zone}",F2:F{last_row-1})')
        z_total.font = BOLD
        z_total.fill = PatternFill("solid", fgColor=ZONE_COLORS[zone])
        z_total.number_format = '#,##0.00 "kr"'
        for c in range(1, 10):
            ws.cell(row=rr, column=c).border = THIN_BORDER
    return r + 3

def write_data_sheet(wb, name, rows):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    write_header(ws)
    color = SHEET_COLORS[name]
    next_r = write_rows(ws, rows, color)
    write_total(ws, next_r, color)
    return ws

# ---------------------------------------------------------------------------
# Zone sheet — one consolidated view per zone. Each section (fundament,
# konstruktion, tag-variant, beklædning-variant) gets a coloured header
# bar, all matching-zone rows (incl. Fælles), and a section subtotal.
# Grand zone-total at the bottom.
# ---------------------------------------------------------------------------
DKK_FMT = '#,##0.00 "kr"'

def write_zone_sheet(wb, sheet_name, zone, sections):
    """Build a consolidated zone sheet.
       sections = list of (label, color_hex, source_rows).
       Filters rows with Zone == `zone` OR Zone == "Fælles"."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    widths = [22, 50, 8, 8, 12, 13, 22, 50, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title bar
    t = ws.cell(row=1, column=1,
                value=f"{zone} — komplet materialeliste (rene {zone}-poster + Fælles)")
    t.font = Font(bold=True, color=HEADER_FG, size=13)
    t.fill = HEADER_FILL
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=9)
    ws.row_dimensions[1].height = 26

    # Column headers
    headers = ["Kategori", "Vare", "Antal", "Enhed", "Pris/enh",
               "I alt", "Leverandør", "Noter", "Zone"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 22
    ws.freeze_panes = "A4"

    r = 4
    subtotal_cells = []
    zone_filter = (zone, "Fælles")

    for label, color, rows in sections:
        # Section header bar
        section_fill = PatternFill("solid", fgColor=color)
        sh = ws.cell(row=r, column=1, value=f"  {label}")
        sh.font = Font(bold=True, size=12)
        sh.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=9)
        for c in range(1, 10):
            ws.cell(row=r, column=c).fill = section_fill
            ws.cell(row=r, column=c).border = THIN_BORDER
        ws.row_dimensions[r].height = 20
        r += 1

        section_start = r
        for row_data in rows:
            kat, vare, antal, enhed, pris, lev, noter, url, rzone = row_data
            if rzone not in zone_filter:
                continue
            a = ws.cell(row=r, column=1, value=kat)
            a.fill = section_fill
            a.font = BOLD
            ws.cell(row=r, column=2, value=vare)
            ws.cell(row=r, column=3, value=antal)
            ws.cell(row=r, column=4, value=enhed).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=5, value=pris).number_format = '#,##0.00'
            if isinstance(antal, (int, float)) and isinstance(pris, (int, float)):
                f = ws.cell(row=r, column=6, value=f"=C{r}*E{r}")
                f.number_format = '#,##0.00'
            lev_cell = ws.cell(row=r, column=7, value=lev)
            if url:
                lev_cell.hyperlink = url
                lev_cell.font = LINK_FONT
            ws.cell(row=r, column=8, value=noter).alignment = Alignment(
                wrap_text=False, vertical="center")
            z = ws.cell(row=r, column=9, value=rzone)
            z.alignment = Alignment(horizontal="center")
            z.fill = PatternFill("solid", fgColor=ZONE_COLORS.get(rzone, "FFFFFF"))
            z.font = BOLD
            for c in range(1, 10):
                ws.cell(row=r, column=c).border = THIN_BORDER
            r += 1
        section_end = r - 1

        # Section subtotal
        stl = ws.cell(row=r, column=5, value=f"{label} subtotal:")
        stl.font = BOLD
        stl.alignment = Alignment(horizontal="right")
        stl.fill = section_fill
        if section_end >= section_start:
            sval = ws.cell(row=r, column=6,
                           value=f"=SUM(F{section_start}:F{section_end})")
            sval.font = BOLD
            sval.fill = TOTAL_FILL
            sval.number_format = DKK_FMT
            subtotal_cells.append(f"F{r}")
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 2  # blank separator before next section

    # Grand total
    r += 1
    gtl = ws.cell(row=r, column=5, value=f"{zone}-ONLY TOTAL:")
    gtl.font = Font(bold=True, size=13)
    gtl.fill = TOTAL_FILL
    gtl.alignment = Alignment(horizontal="right")
    if subtotal_cells:
        gt = ws.cell(row=r, column=6, value=f"={'+'.join(subtotal_cells)}")
        gt.font = Font(bold=True, size=13)
        gt.fill = TOTAL_FILL
        gt.number_format = DKK_FMT
    for c in range(1, 10):
        ws.cell(row=r, column=c).border = THIN_BORDER
    ws.row_dimensions[r].height = 26

    return ws

# ---------------------------------------------------------------------------
# Cladding price comparison sheet — all options as rows, sorted by total.
# Net cladded wall area: ~16 m² (7,81 m perimeter × ~2,3 m gns højde minus
# vinduer + døre). Antal beregnet pr. produkt-mønster-kombo.
# ---------------------------------------------------------------------------

# Each row: dict med produkt, mønster, materiale, antal, enhed, pris_stk,
# behandling_kr, leverandør, url, note.
CLADDING_OPTIONS = [
    {
        "produkt":    "Imp. klinkbeklædning gran 25×125×3600",
        "monster":    "Klink horisontal",
        "materiale":  "Trykimp. gran (ubehandlet ud over imp.)",
        "antal":      43, "enhed": "stk",
        "kr_stk":     60.30,
        "behandling": 0,   # imp. behøver ikke maling; kan stå råt
        "leverandor": "jemogfix.dk",
        "url":        "https://www.jemogfix.dk/impraegneret-klinkbeklaedning-gran-25-x-125-x-3600-mm/",
        "note":       "Billigste klink. Grøn-grålig fra start; kan males sort senere (+~450 kr)",
    },
    {
        "produkt":    "Aros 1-på-2 svensk gran 25×150 savskåret",
        "monster":    "1-på-2 lodret (bob)",
        "materiale":  "Savskåret svensk gran (FSC)",
        "antal":      16, "enhed": "m²",
        "kr_stk":     160.00,
        "behandling": 450,   # 3 L Pinotex/Gori grunding + topcoat
        "leverandor": "arossavvaerk.dk",
        "url":        "https://arossavvaerk.dk/vare/1-paa-2-beklaedning-svensk-gran/",
        "note":       "Tilbudspris 160 (normal 225). Rustic look. Behandling påkrævet før montage",
    },
    {
        "produkt":    "Klinkbeklædning sortmalet gran 25×125×4200 (6-pak)",
        "monster":    "Klink horisontal",
        "materiale":  "Gran, færdig sortmalet vandbaseret RAL 9005",
        "antal":      6, "enhed": "pak",
        "kr_stk":     497.70,
        "behandling": 0,    # færdig sortmalet fra start
        "leverandor": "jemogfix.dk",
        "url":        "https://www.jemogfix.dk/klinkbeklaedning-sortmalet/4115/9023155/",
        "note":       "Klar til at montere. Klassisk skur/sommerhus-look. 6 stk pr. pak = 25,2 m linear",
    },
    {
        "produkt":    "Høvlet forskalling fyr 22×100×3600 UBEHANDLET",
        "monster":    "1-på-2 lodret (bob)",
        "materiale":  "Fyr, høvlet, ubehandlet",
        "antal":      120, "enhed": "stk",
        "kr_stk":     35.82,
        "behandling": 600,   # intensiv behandling kræves
        "leverandor": "jemogfix.dk",
        "url":        "https://www.jemogfix.dk/hoevlet-forskalling-22-x-100-x-3600-mm/4230/9057349/",
        "note":       "Jemogfix advarer mod udendørs ubehandlet. Kræver Gori/Pinotex × 3 lag",
    },
    {
        "produkt":    "Beklædning lodret sort 25×125×3000 (4-pak)",
        "monster":    "Klink lodret",
        "materiale":  "Skandinavisk gran (FSC), færdig sortmalet vandbaseret",
        "antal":      13, "enhed": "pak",
        "kr_stk":     359.40,
        "behandling": 0,
        "leverandor": "jemogfix.dk",
        "url":        "https://www.jemogfix.dk/beklaedning-lodret-sort-25-x-125-x-3000-mm-4-pk/4115/9064083/",
        "note":       "Lodret klink-profil, færdig sort. 4 stk pr. pak = 12 m linear. Moderne lade-look",
    },
    {
        "produkt":    "Imprægneret høvlet fyr 25×125×3600",
        "monster":    "1-på-2 lodret (bob)",
        "materiale":  "Trykimp. fyr, høvlet",
        "antal":      81, "enhed": "stk",
        "kr_stk":     63.90,
        "behandling": 0,
        "leverandor": "jemogfix.dk",
        "url":        "https://www.jemogfix.dk/impraegnerede-braedder-fyr-25-x-125-x-3600-mm/",
        "note":       "Udendørs-godkendt. Grøn-brun farve fra imp. Kan males (+~450 kr)",
    },
]

def write_cladding_comparison(wb):
    name = "sammenligning_beklaedning"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    # Title bar
    t = ws.cell(row=1, column=1,
                value="Beklædning — prissammenligning ved 16 m² netto facadeareal")
    t.font = Font(bold=True, color=HEADER_FG, size=13)
    t.fill = HEADER_FILL
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:L1")
    ws.row_dimensions[1].height = 24

    # Column headers
    headers = [
        ("Produkt", 50), ("Mønster", 18), ("Materiale", 32),
        ("Antal", 8), ("Enhed", 8), ("Pris/enh", 11),
        ("Material kr", 13), ("Behandl. kr", 12), ("TOTAL", 13),
        ("Δ vs billigst", 14), ("Leverandør", 16), ("Note", 60),
    ]
    for i, (h, w) in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 20

    # Sort options by total ascending
    sorted_opts = sorted(
        CLADDING_OPTIONS,
        key=lambda o: o["antal"] * o["kr_stk"] + o["behandling"]
    )

    # Pattern → background color
    pat_color = {
        "Klink horisontal":     COL_CLAD_KLINK,
        "Klink lodret":         COL_TAG_ETERN,
        "1-på-2 lodret (bob)":  COL_CLAD_BOB,
    }

    min_total = min(o["antal"] * o["kr_stk"] + o["behandling"] for o in sorted_opts)

    DKK = '#,##0.00 "kr"'
    DKK_INT = '#,##0 "kr"'

    for i, o in enumerate(sorted_opts):
        r = 4 + i
        total = o["antal"] * o["kr_stk"] + o["behandling"]
        delta = total - min_total

        ws.cell(row=r, column=1, value=o["produkt"])
        b = ws.cell(row=r, column=2, value=o["monster"])
        b.fill = PatternFill("solid", fgColor=pat_color[o["monster"]])
        b.font = BOLD
        b.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value=o["materiale"])
        ws.cell(row=r, column=4, value=o["antal"]).alignment = Alignment(horizontal="right")
        ws.cell(row=r, column=5, value=o["enhed"]).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=6, value=o["kr_stk"]).number_format = DKK
        ws.cell(row=r, column=7, value=f"=D{r}*F{r}").number_format = DKK
        ws.cell(row=r, column=8, value=o["behandling"]).number_format = DKK_INT
        tot = ws.cell(row=r, column=9, value=f"=G{r}+H{r}")
        tot.number_format = DKK
        tot.font = BOLD
        if i == 0:
            tot.fill = TOTAL_FILL   # billigste = guld
        dlt = ws.cell(row=r, column=10, value=delta if delta > 0 else None)
        dlt.number_format = DKK
        dlt.alignment = Alignment(horizontal="right")
        if delta > 0:
            dlt.fill = DELTA_POS_FILL

        lev = ws.cell(row=r, column=11, value=o["leverandor"])
        if o.get("url"):
            lev.hyperlink = o["url"]
            lev.font = LINK_FONT
        ws.cell(row=r, column=12, value=o["note"]).alignment = Alignment(wrap_text=False, vertical="center")

        for c in range(1, 13):
            ws.cell(row=r, column=c).border = THIN_BORDER

    # Legend block
    legend_row = 4 + len(sorted_opts) + 2
    ws.cell(row=legend_row, column=1, value="Læsevejledning").font = HEADER_FONT
    ws.cell(row=legend_row, column=1).fill = SUBHEADER_FILL
    ws.merge_cells(start_row=legend_row, end_row=legend_row, start_column=1, end_column=6)

    notes = [
        "Antal er beregnet for 16 m² netto vægareal (4 hus-vægge minus side-vindue + dør + kanindør).",
        "Behandl. kr = ekstra behandlingsomkostning (træolie/maling) der KRÆVES før montage. 0 hvis produktet er færdigbehandlet eller trykimp.",
        "Klink-produkter har færdig profil (overlapper indbygget). 1-på-2 (bob) bruger to lag rå brædder; kræver mere arbejde at montere.",
        "Mønster-farven matcher beklædnings-sheet i resten af workbook'en: blå = klink, grøn = bob, gul = lodret klink.",
        "TOTAL inkluderer IKKE: voliernet, vindpap, klemmelister, isolering, hjørnetrim (~2.000-2.500 kr fælles, se beklaedning_* sheets).",
    ]
    for i, n in enumerate(notes):
        c = ws.cell(row=legend_row + 1 + i, column=1, value=f"•  {n}")
        ws.merge_cells(start_row=legend_row+1+i, end_row=legend_row+1+i,
                       start_column=1, end_column=12)
        c.alignment = Alignment(wrap_text=True, vertical="top")

    return ws


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------
def write_summary(wb):
    if "summary" in wb.sheetnames:
        del wb["summary"]
    ws = wb.create_sheet("summary", 0)   # first sheet

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 26

    DKK_FMT = '#,##0.00 "kr"'

    # --- title bar ---
    t = ws.cell(row=1, column=1, value="Rabbit-house BOM  —  variant selector")
    t.font = Font(bold=True, color=HEADER_FG, size=14)
    t.fill = HEADER_FILL
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 28

    # --- selector block ---
    ws.cell(row=3, column=1, value="Vælg tagdækning:").font = BOLD
    sel_roof = ws.cell(row=3, column=2, value="tagpap")
    sel_roof.fill = PatternFill("solid", fgColor=COL_TAG_TAGPAP)
    sel_roof.font = BOLD
    sel_roof.alignment = Alignment(horizontal="center")
    dv_roof = DataValidation(type="list", formula1='"tagpap,eternit_b7"', allow_blank=False)
    dv_roof.add("B3")
    ws.add_data_validation(dv_roof)

    ws.cell(row=4, column=1, value="Vælg beklædning:").font = BOLD
    sel_clad = ws.cell(row=4, column=2, value="klink")
    sel_clad.fill = PatternFill("solid", fgColor=COL_CLAD_KLINK)
    sel_clad.font = BOLD
    sel_clad.alignment = Alignment(horizontal="center")
    dv_clad = DataValidation(type="list", formula1='"klink,board_on_board"', allow_blank=False)
    dv_clad.add("B4")
    ws.add_data_validation(dv_clad)

    # --- section header: Selected totals ---
    h = ws.cell(row=6, column=1, value="Valgt konfiguration — totaler")
    h.font = HEADER_FONT
    h.fill = SUBHEADER_FILL
    ws.merge_cells("A6:D6")
    ws.row_dimensions[6].height = 20

    # Per-category rows. Column A gets the category color, B the value.
    sel_rows = [
        ("Fundament",      COL_FUNDAMENT,  "=SUM(fundament!F:F)"),
        ("Konstruktion",   COL_KONSTRUKT,  "=SUM(konstruktion!F:F)"),
        ("Tagkonstruktion","TAG_DYNAMIC",  '=SUM(INDIRECT("tagkonstruktion_"&B3&"!F:F"))'),
        ("Beklædning",     "CLAD_DYNAMIC", '=SUM(INDIRECT("beklaedning_"&B4&"!F:F"))'),
    ]
    for i, (label, color, formula) in enumerate(sel_rows):
        r = 7 + i
        a = ws.cell(row=r, column=1, value=label)
        a.font = BOLD
        if color == "TAG_DYNAMIC":
            # No fixed color — formula based, leave neutral light gray
            a.fill = PatternFill("solid", fgColor="F2F2F2")
        elif color == "CLAD_DYNAMIC":
            a.fill = PatternFill("solid", fgColor="F2F2F2")
        else:
            a.fill = PatternFill("solid", fgColor=color)
        b = ws.cell(row=r, column=2, value=formula)
        b.number_format = DKK_FMT
        b.alignment = Alignment(horizontal="right")
        a.border = THIN_BORDER
        b.border = THIN_BORDER

    # Grand total row
    r = 11
    a = ws.cell(row=r, column=1, value="TOTAL valgt")
    a.font = Font(bold=True, size=12)
    a.fill = TOTAL_FILL
    a.alignment = Alignment(horizontal="right")
    b = ws.cell(row=r, column=2, value="=SUM(B7:B10)")
    b.font = Font(bold=True, size=12)
    b.fill = TOTAL_FILL
    b.number_format = DKK_FMT
    b.alignment = Alignment(horizontal="right")
    a.border = THIN_BORDER
    b.border = THIN_BORDER

    # --- section header: Comparison matrix ---
    h2 = ws.cell(row=13, column=1, value="Sammenligning — alle 4 kombinationer")
    h2.font = HEADER_FONT
    h2.fill = SUBHEADER_FILL
    ws.merge_cells("A13:D13")
    ws.row_dimensions[13].height = 20

    # Matrix column headers
    c1 = ws.cell(row=14, column=2, value="tagpap")
    c1.fill = PatternFill("solid", fgColor=COL_TAG_TAGPAP); c1.font = BOLD
    c1.alignment = Alignment(horizontal="center")
    c2 = ws.cell(row=14, column=3, value="eternit_b7")
    c2.fill = PatternFill("solid", fgColor=COL_TAG_ETERN);  c2.font = BOLD
    c2.alignment = Alignment(horizontal="center")
    c3 = ws.cell(row=14, column=4, value="Δ (eternit − tagpap)")
    c3.font = BOLD
    c3.alignment = Alignment(horizontal="center")
    for col in range(2, 5):
        ws.cell(row=14, column=col).border = THIN_BORDER

    # Klink row
    k = ws.cell(row=15, column=1, value="Klink")
    k.fill = PatternFill("solid", fgColor=COL_CLAD_KLINK); k.font = BOLD
    k.alignment = Alignment(horizontal="right")
    ws.cell(row=15, column=2, value="=SUM(fundament!F:F)+SUM(konstruktion!F:F)+SUM(tagkonstruktion_tagpap!F:F)+SUM(beklaedning_klink!F:F)")
    ws.cell(row=15, column=3, value="=SUM(fundament!F:F)+SUM(konstruktion!F:F)+SUM(tagkonstruktion_eternit_b7!F:F)+SUM(beklaedning_klink!F:F)")
    ws.cell(row=15, column=4, value="=C15-B15")
    ws.cell(row=15, column=4).fill = DELTA_POS_FILL

    # Bob row
    b = ws.cell(row=16, column=1, value="Board-on-board (1-på-2)")
    b.fill = PatternFill("solid", fgColor=COL_CLAD_BOB); b.font = BOLD
    b.alignment = Alignment(horizontal="right")
    ws.cell(row=16, column=2, value="=SUM(fundament!F:F)+SUM(konstruktion!F:F)+SUM(tagkonstruktion_tagpap!F:F)+SUM(beklaedning_board_on_board!F:F)")
    ws.cell(row=16, column=3, value="=SUM(fundament!F:F)+SUM(konstruktion!F:F)+SUM(tagkonstruktion_eternit_b7!F:F)+SUM(beklaedning_board_on_board!F:F)")
    ws.cell(row=16, column=4, value="=C16-B16")
    ws.cell(row=16, column=4).fill = DELTA_POS_FILL

    # Delta row (bob - klink)
    d = ws.cell(row=17, column=1, value="Δ (bob − klink)")
    d.font = BOLD
    d.alignment = Alignment(horizontal="right")
    d.fill = PatternFill("solid", fgColor="F2F2F2")
    ws.cell(row=17, column=2, value="=B16-B15").fill = DELTA_POS_FILL
    ws.cell(row=17, column=3, value="=C16-C15").fill = DELTA_POS_FILL

    # Borders + number format on entire matrix
    for r in range(15, 18):
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            if c in (2, 3, 4) and cell.value is not None:
                cell.number_format = DKK_FMT
                cell.alignment = Alignment(horizontal="right")

    # --- notes ---
    n = ws.cell(row=19, column=1, value="Noter")
    n.font = HEADER_FONT
    n.fill = SUBHEADER_FILL
    ws.merge_cells("A19:D19")

    notes = [
        "Dropdowns i B3/B4 styrer totalen i B11 via INDIRECT på sheet-navne.",
        "Eternit_b7 bruger tungere spær (47×100 C24) end tagpap (men prisen er fra eternit-doc og er et estimat).",
        "Board-on-board pris er Aros tilbudspris 160 kr/m² (normalpris 225). Klink kommer sortmalet; bob behøver træolie/maling.",
        "Voliernet og isolering er ens på tværs af beklædningsvarianter.",
        "Variantvalg i build.scad: roof_cover = \"tagpap_osb\"|\"eternit_b7\"; cladding_type = \"klink\"|\"board_on_board\".",
        "Items med pris=tom og noter=\"TBD …\" kræver verifikation før bestilling.",
    ]
    for i, note in enumerate(notes):
        ws.cell(row=20 + i, column=1, value=f"•  {note}")
        ws.merge_cells(start_row=20+i, end_row=20+i, start_column=1, end_column=4)
        ws.cell(row=20+i, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    return ws

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    wb = openpyxl.load_workbook(XLSX)

    # 1. drop scratch sheets
    for stale in ("tagkonstruktion_tagpap", "materialeliste"):
        if stale in wb.sheetnames:
            del wb[stale]

    # 2. write new data sheets
    write_data_sheet(wb, "fundament",                  FUNDAMENT)
    write_data_sheet(wb, "konstruktion",               KONSTRUKTION)
    write_data_sheet(wb, "tagkonstruktion_tagpap",     TAG_TAGPAP)
    write_data_sheet(wb, "tagkonstruktion_eternit_b7", TAG_ETERNIT)
    write_data_sheet(wb, "beklaedning_klink",          CLAD_KLINK)
    write_data_sheet(wb, "beklaedning_board_on_board", CLAD_BOB)

    # 3. summary + cladding comparison
    write_summary(wb)
    write_cladding_comparison(wb)

    # 4. zone sheets — consolidated view per zone (default variant: eternit_b7
    # + klink, matching main.scad). Edit the section lists if you switch
    # variant and want the zone view to follow.
    hus_sections = [
        ("Fundament",                     COL_FUNDAMENT,  FUNDAMENT),
        ("Konstruktion",                  COL_KONSTRUKT,  KONSTRUKTION),
        ("Tagkonstruktion (eternit_b7)",  COL_TAG_ETERN,  TAG_ETERNIT),
        ("Beklædning (klink)",            COL_CLAD_KLINK, CLAD_KLINK),
    ]
    write_zone_sheet(wb, "hus", "Hus", hus_sections)

    yard_sections = [
        ("Fundament",                     COL_FUNDAMENT,  FUNDAMENT),
        ("Konstruktion",                  COL_KONSTRUKT,  KONSTRUKTION),
        ("Tagkonstruktion (eternit_b7)",  COL_TAG_ETERN,  TAG_ETERNIT),
        ("Beklædning / voliere (klink)",  COL_CLAD_KLINK, CLAD_KLINK),
    ]
    write_zone_sheet(wb, "yard", "Yard", yard_sections)

    # 5. order: summary, zone overviews, comparison, common, then variants
    order = [
        "summary",
        "hus",
        "yard",
        "sammenligning_beklaedning",
        "fundament",
        "konstruktion",
        "tagkonstruktion_tagpap",
        "tagkonstruktion_eternit_b7",
        "beklaedning_klink",
        "beklaedning_board_on_board",
    ]
    wb._sheets = [wb[n] for n in order]

    # Format number columns in data sheets (skip zone sheets — handled in writer)
    skip_zone_format = {"hus", "yard", "summary", "sammenligning_beklaedning"}
    for sheet_name in order:
        if sheet_name in skip_zone_format:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2):
            # E = pris, F = i alt
            row[4].number_format = '#,##0.00'   # Pris
            row[5].number_format = '#,##0.00'   # I alt

    wb.save(XLSX)
    print(f"Wrote {XLSX}")
    print(f"Sheets: {wb.sheetnames}")

if __name__ == "__main__":
    main()
